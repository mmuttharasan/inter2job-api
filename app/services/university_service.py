"""
University Admin Service Layer — business logic for /api/universities/* endpoints.

University admins manage their own institution: profile, departments, and student onboarding.
"""

import io
import secrets
import string
import uuid
from datetime import datetime, timezone

from ..services.supabase_client import supabase


# ═══════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════

def _now_iso() -> str:
    return datetime.now(tz=timezone.utc).isoformat()


def _generate_temp_password(length: int = 12) -> str:
    """Generate a cryptographically secure temporary password."""
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _get_university_id_for_user(user_id: str) -> str | None:
    """Return the university_id from the user's profile row, or None."""
    try:
        res = (
            supabase.table("profiles")
            .select("university_id")
            .eq("id", user_id)
            .single()
            .execute()
        )
        return (res.data or {}).get("university_id")
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════════════
# §1  University Profile
# ═══════════════════════════════════════════════════════════════════════════

def get_my_university(user_id: str) -> dict:
    """Fetch the university linked to a university_admin user."""
    university_id = _get_university_id_for_user(user_id)
    if not university_id:
        return {}
    try:
        res = (
            supabase.table("universities")
            .select("id, name, domain, logo_url, location, status, created_at, updated_at")
            .eq("id", university_id)
            .single()
            .execute()
        )
        return res.data or {}
    except Exception:
        return {}


def update_my_university(university_id: str, data: dict) -> dict:
    """Update allowed university profile fields."""
    allowed = {"name", "location", "logo_url", "domain"}
    update = {k: v for k, v in data.items() if k in allowed and v is not None}
    if not update:
        raise ValueError("NO_CHANGES")
    update["updated_at"] = _now_iso()
    try:
        res = (
            supabase.table("universities")
            .update(update)
            .eq("id", university_id)
            .execute()
        )
        rows = res.data or []
        return rows[0] if rows else {}
    except Exception as e:
        raise ValueError(str(e))


# ═══════════════════════════════════════════════════════════════════════════
# §2  Departments
# ═══════════════════════════════════════════════════════════════════════════

def list_my_departments(university_id: str) -> list:
    """List all departments for the university."""
    try:
        res = (
            supabase.table("university_departments")
            .select(
                "id, university_id, name, code, head, students_count, placed_count, "
                "faculty_count, labs_count, avg_package, created_at, updated_at"
            )
            .eq("university_id", university_id)
            .order("name")
            .execute()
        )
        return res.data or []
    except Exception:
        return []


def create_my_department(university_id: str, actor_id: str, data: dict) -> dict:
    """Create a new department for the university."""
    name = (data.get("name") or "").strip()
    code = (data.get("code") or "").strip()
    if not name or not code:
        raise ValueError("MISSING_FIELDS")

    dept = {
        "id": str(uuid.uuid4()),
        "university_id": university_id,
        "name": name,
        "code": code.upper(),
        "head": data.get("head"),
        "students_count": int(data.get("students_count") or 0),
        "placed_count": int(data.get("placed_count") or 0),
        "faculty_count": int(data.get("faculty_count") or 0),
        "labs_count": int(data.get("labs_count") or 0),
        "avg_package": data.get("avg_package"),
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    try:
        res = supabase.table("university_departments").insert(dept).execute()
        rows = res.data or []
        return rows[0] if rows else dept
    except Exception as e:
        err = str(e).lower()
        if "duplicate" in err or "unique" in err:
            raise ValueError("DEPT_EXISTS")
        raise ValueError(str(e))


def update_my_department(university_id: str, dept_id: str, data: dict) -> dict:
    """Update a department. Only allows updating own university's departments."""
    allowed = {"name", "code", "head", "students_count", "placed_count",
               "faculty_count", "labs_count", "avg_package"}
    update = {k: v for k, v in data.items() if k in allowed and v is not None}
    if not update:
        raise ValueError("NO_CHANGES")
    update["updated_at"] = _now_iso()
    try:
        res = (
            supabase.table("university_departments")
            .update(update)
            .eq("id", dept_id)
            .eq("university_id", university_id)
            .execute()
        )
        rows = res.data or []
        if not rows:
            raise ValueError("NOT_FOUND")
        return rows[0]
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(str(e))


# ═══════════════════════════════════════════════════════════════════════════
# §3  Students
# ═══════════════════════════════════════════════════════════════════════════

def list_my_students(university_id: str, params: dict) -> dict:
    """List students belonging to the university with optional filtering."""
    page = max(1, int(params.get("page") or 1))
    limit = min(200, max(1, int(params.get("limit") or 50)))
    offset = (page - 1) * limit
    search = (params.get("search") or "").strip().lower()
    department = (params.get("department") or "").strip()

    try:
        query = (
            supabase.table("students")
            .select(
                "id, university_id, department, graduation_year, gpa, "
                "verification_status, created_at, profiles(full_name, avatar_url)"
            )
            .eq("university_id", university_id)
        )
        if department:
            query = query.eq("department", department)

        res = query.order("created_at", desc=True).execute()
        all_rows = res.data or []
    except Exception:
        all_rows = []

    # Flatten profiles join and apply name search in Python
    result = []
    for s in all_rows:
        profile = s.pop("profiles", None) or {}
        s["full_name"] = profile.get("full_name") or ""
        s["avatar_url"] = profile.get("avatar_url")
        if search and search not in s["full_name"].lower():
            continue
        result.append(s)

    total = len(result)
    paginated = result[offset: offset + limit]

    return {
        "students": paginated,
        "page": page,
        "limit": limit,
        "total": total,
    }


def create_student(university_id: str, actor_id: str, data: dict) -> dict:
    """Create a single student auth account and link to the university."""
    email = (data.get("email") or "").strip()
    full_name = (data.get("full_name") or "").strip()
    password = (data.get("password") or "").strip() or _generate_temp_password()

    if not email or not full_name:
        raise ValueError("MISSING_FIELDS")

    # 1. Create Supabase auth user
    try:
        response = supabase.auth.admin.create_user({
            "email": email,
            "password": password,
            "email_confirm": True,
            "user_metadata": {"full_name": full_name, "role": "student"},
        })
        user_id = response.user.id
    except Exception as e:
        err = str(e).lower()
        if "already" in err or "exists" in err or "duplicate" in err:
            raise ValueError("EMAIL_EXISTS")
        raise ValueError(f"CREATE_FAILED:{e}")

    # 2. Upsert profile
    supabase.table("profiles").upsert({
        "id": user_id,
        "role": "student",
        "full_name": full_name,
        "university_id": university_id,
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }).execute()

    # 3. Upsert student record
    student_row: dict = {
        "id": user_id,
        "university_id": university_id,
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    if data.get("department"):
        student_row["department"] = data["department"]
    if data.get("graduation_year"):
        student_row["graduation_year"] = int(data["graduation_year"])
    if data.get("gpa"):
        try:
            student_row["gpa"] = float(data["gpa"])
        except (ValueError, TypeError):
            pass

    supabase.table("students").upsert(student_row).execute()

    return {
        "id": user_id,
        "email": email,
        "full_name": full_name,
        "university_id": university_id,
        "department": data.get("department"),
        "graduation_year": data.get("graduation_year"),
        "temp_password": password if not data.get("password") else None,
    }


def get_university_student_detail(university_id: str, student_id: str) -> dict | None:
    """Return comprehensive student data for university admin view."""

    # 1. Verify student belongs to this university
    try:
        sr = (
            supabase.table("students")
            .select("*")
            .eq("id", student_id)
            .eq("university_id", university_id)
            .maybe_single()
            .execute()
        )
        if not sr.data:
            return None
        student = sr.data
    except Exception:
        return None

    # 2. Profile
    profile: dict = {}
    try:
        pr = (
            supabase.table("profiles")
            .select("full_name, avatar_url")
            .eq("id", student_id)
            .maybe_single()
            .execute()
        )
        profile = pr.data or {}
    except Exception:
        pass

    # 3. Applications → jobs → companies
    applications: list = []
    try:
        ar = (
            supabase.table("applications")
            .select("status, applied_at, shortlisted_at, jobs(id, title, companies(name, logo_url))")
            .eq("student_id", student_id)
            .execute()
        )
        for app in (ar.data or []):
            job = app.get("jobs") or {}
            company = job.get("companies") or {}
            applications.append({
                "job_title": job.get("title", ""),
                "company_name": company.get("name", ""),
                "company_logo_url": company.get("logo_url"),
                "status": app.get("status", "pending"),
                "applied_at": app.get("applied_at") or app.get("shortlisted_at"),
            })
    except Exception:
        pass

    # 4. Internships → jobs + companies
    internships: list = []
    try:
        ir = (
            supabase.table("internships")
            .select("id, status, start_date, end_date, jobs(title), companies(name, logo_url)")
            .eq("student_id", student_id)
            .order("created_at", desc=True)
            .execute()
        )
        for intern in (ir.data or []):
            internships.append({
                "id": intern.get("id"),
                "job_title": (intern.get("jobs") or {}).get("title", ""),
                "company_name": (intern.get("companies") or {}).get("name", ""),
                "company_logo_url": (intern.get("companies") or {}).get("logo_url"),
                "status": intern.get("status", ""),
                "start_date": intern.get("start_date"),
                "end_date": intern.get("end_date"),
            })
    except Exception:
        pass

    # 5. Job fit: find top matching published jobs by skill overlap
    skills = student.get("skills") or []
    job_fits: list = []
    try:
        jr = (
            supabase.table("jobs")
            .select("id, title, skills, companies(name)")
            .eq("status", "published")
            .execute()
        )
        student_skill_set = {s.lower() for s in skills}
        for job in (jr.data or [])[:100]:
            job_skills = job.get("skills") or []
            if not job_skills or not student_skill_set:
                continue
            job_set = {s.lower() for s in job_skills}
            matched = student_skill_set & job_set
            if not matched:
                continue
            score = round(len(matched) / len(job_set) * 100)
            if score < 25:
                continue
            job_fits.append({
                "job_id": job["id"],
                "title": job["title"],
                "company_name": ((job.get("companies") or {}).get("name") or ""),
                "match_score": score,
                "matched_skills": list(matched)[:5],
            })
        job_fits = sorted(job_fits, key=lambda x: x["match_score"], reverse=True)[:5]
    except Exception:
        pass

    offers_count = sum(1 for a in applications if a.get("status") == "offered")
    active_internship = next((i for i in internships if i.get("status") == "active"), None)
    interview_apps = [a for a in applications if a.get("status") in ("shortlisted", "interview", "interviewing")]

    return {
        "id": student_id,
        "full_name": profile.get("full_name") or "",
        "avatar_url": profile.get("avatar_url"),
        "department": student.get("department"),
        "graduation_year": student.get("graduation_year"),
        "gpa": str(student["gpa"]) if student.get("gpa") is not None else None,
        "skills": skills,
        "jp_level": student.get("jp_level"),
        "bio": student.get("bio"),
        "linkedin": student.get("linkedin"),
        "github": student.get("github"),
        "verification_status": student.get("verification_status", "unverified"),
        "profile_completeness": float(student.get("profile_completeness") or 0),
        "applications": applications,
        "internships": internships,
        "job_fits": job_fits,
        "offers_count": offers_count,
        "active_internship": active_internship,
        "in_interview": len(interview_apps) > 0,
        "interview_count": len(interview_apps),
    }


def bulk_create_students(university_id: str, actor_id: str, students: list) -> dict:
    """Bulk-create student accounts for the university."""
    created = []
    errors = []

    for i, student_data in enumerate(students):
        try:
            result = create_student(university_id, actor_id, student_data)
            created.append(result)
        except ValueError as e:
            errors.append({
                "row": i + 1,
                "email": student_data.get("email", ""),
                "error": str(e),
            })
        except Exception:
            errors.append({
                "row": i + 1,
                "email": student_data.get("email", ""),
                "error": "UNEXPECTED_ERROR",
            })

    return {
        "created": created,
        "errors": errors,
        "total": len(students),
        "success_count": len(created),
        "error_count": len(errors),
    }


# ═══════════════════════════════════════════════════════════════════════════
# §4  Excel Template
# ═══════════════════════════════════════════════════════════════════════════

def build_student_template() -> io.BytesIO:
    """Generate an .xlsx template file for bulk student import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["full_name", "email", "password", "department", "graduation_year", "gpa"]
    notes = [
        "Full name (required)",
        "Email address (required)",
        "Temp password (leave blank to auto-generate)",
        "Department code e.g. CSE, ECE",
        "Graduation year e.g. 2025",
        "GPA / CGPA e.g. 8.5",
    ]
    example = ["Arjun Ramesh", "arjun@anna.edu", "", "CSE", 2025, 8.9]

    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    note_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    example_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")

    for col, (header, note, val) in enumerate(zip(headers, notes, example), start=1):
        # Row 1 — column name
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Row 2 — helper note
        note_cell = ws.cell(row=2, column=col, value=note)
        note_cell.font = Font(italic=True, color="64748B", size=10)
        note_cell.fill = note_fill
        # Row 3 — example row (yellow, skipped during import)
        ex_cell = ws.cell(row=3, column=col, value=val)
        ex_cell.fill = example_fill
        ex_cell.font = Font(italic=True, color="92400E", size=10)

    # Label row 3 clearly
    ws.cell(row=3, column=1, value="EXAMPLE — Arjun Ramesh (this row is skipped during import)")

    # Row 4+ — blank rows ready for user data
    widths = [30, 32, 22, 15, 18, 10]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_student_excel(stream) -> list[dict]:
    """Parse an uploaded .xlsx file into a list of student dicts."""
    import openpyxl

    wb = openpyxl.load_workbook(stream)
    ws = wb.active

    # Row 1 = headers, Row 2 = notes, Row 3+ = data
    raw_headers = [
        str(ws.cell(row=1, column=col).value or "").strip().lower()
        for col in range(1, ws.max_column + 1)
    ]

    students = []
    # min_row=4: row 1 = headers, row 2 = notes, row 3 = example (skip), row 4+ = data
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        row_data = dict(zip(raw_headers, row))

        full_name = str(row_data.get("full_name") or "").strip()
        email = str(row_data.get("email") or "").strip()
        if not full_name or not email:
            continue

        # graduation_year: Excel returns floats (2025.0), handle with float→int
        gy_raw = row_data.get("graduation_year")
        try:
            graduation_year = int(float(str(gy_raw).strip())) if gy_raw is not None else None
        except (ValueError, TypeError):
            graduation_year = None

        # gpa: wrap individually so one bad cell doesn't abort the whole file
        gpa_raw = row_data.get("gpa")
        try:
            gpa = float(gpa_raw) if gpa_raw is not None else None
        except (ValueError, TypeError):
            gpa = None

        students.append({
            "full_name": full_name,
            "email": email,
            "password": str(row_data.get("password") or "").strip() or None,
            "department": str(row_data.get("department") or "").strip() or None,
            "graduation_year": graduation_year,
            "gpa": gpa,
        })

    return students
